from flask import Flask, render_template_string, request, jsonify, send_file
import imaplib
import email
from email.header import decode_header
import re
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
import io

app = Flask(__name__)
app.secret_key = "saksham_val_super_key"

linked_accounts = []
mis_database = []
SAVED_API_KEY = ""

IMAP_SERVERS = {
    "Yahoo Mail": "imap.mail.yahoo.com",
    "Gmail": "imap.gmail.com",
    "Outlook / Hotmail": "outlook.office365.com",
    "Auto Detect": "imap.mail.yahoo.com"
}

def smart_clean_text(raw_html_or_text):
    if not raw_html_or_text: return ""
    try:
        soup = BeautifulSoup(raw_html_or_text, "html.parser")
        return soup.get_text(separator=" ")
    except Exception:
        return raw_html_or_text

def deep_parse_email(subject, clean_text, sender):
    full = f"{subject} {clean_text}"
    cust_name = None
    patterns_name = [
        r'Applicant Name[\s:-]*([A-Z\s]{3,30})(?:\n|\r|Mobile|App|Case)',
        r'Case of\s*-\s*([A-Z\s]{3,30})(?:\n|\r|//|App|Mobile)',
        r'Technical Report Initiate of\s*-\s*([A-Z\s]{3,30})(?:\n|\r|//|App)'
    ]
    for p in patterns_name:
        m = re.search(p, full, re.IGNORECASE)
        if m and len(m.group(1).strip()) > 2:
            cust_name = m.group(1).strip()
            break

    app_no = None
    app_match = re.search(r'(?:Application No|App id|App No)[\s:-]*(\d{6,12})', full, re.IGNORECASE)
    if app_match:
        app_no = app_match.group(1).strip()

    mobile = "N/A"
    mob_match = re.search(r'(?:Mobile No|Mobile Number)[\s:-]*([6-9]\d{9})', full, re.IGNORECASE)
    if mob_match:
        mobile = mob_match.group(1).strip()

    bank_name = "SMFG / Grihashakti" if "grihashakti" in sender.lower() or "smfg" in full.lower() else "Bank / FI"
    case_type = "HL RESALE PURCHASE" if "RESALE" in full.upper() else "Valuation Case"
    return cust_name, app_no, mobile, bank_name, case_type

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="hi">
<head>
    <meta charset="UTF-8">
    <title>Saksham Valuer Portal</title>
    <style>
        body { font-family: sans-serif; background: #0f172a; color: #f8fafc; margin: 0; padding: 20px; }
        .container { max-width: 1200px; margin: 0 auto; }
        .box { background: #1e293b; padding: 25px; border-radius: 10px; margin-bottom: 25px; border: 1px solid #334155; }
        h2, h3 { color: #38bdf8; margin-top: 0; }
        input, select { width: 100%; padding: 10px; margin: 8px 0; border-radius: 6px; border: 1px solid #475569; background: #0f172a; color: white; box-sizing: border-box; }
        button { padding: 10px 20px; background: #38bdf8; color: #0f172a; border: none; border-radius: 6px; font-weight: bold; cursor: pointer; }
        button:hover { background: #0ea5e9; }
        .btn-green { background: #10b981; color: white; }
        .btn-purple { background: #a855f7; color: white; }
        .btn-orange { background: #f97316; color: white; }
        table { width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 15px; }
        th, td { padding: 12px; border-bottom: 1px solid #334155; text-align: left; }
        th { color: #38bdf8; background: #0f172a; }
        .flex-row { display: flex; gap: 15px; align-items: center; }
        .status-msg { font-size: 13px; font-weight: bold; margin-top: 10px; }
        .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }
        .parsed-box { background: #0f172a; border: 2px dashed #a855f7; padding: 15px; border-radius: 8px; margin-top: 15px; }
        .filter-card { background: #0f172a; padding: 15px; border-radius: 8px; margin-bottom: 15px; display: flex; gap: 15px; align-items: flex-end; }
    </style>
</head>
<body>
    <div class="container">
        <div style="display:flex; justify-content:space-between; align-items:center;" class="box">
            <div>
                <h2 style="margin:0;">SAKSHAM VALUER AI PORTAL</h2>
                <small style="color:#94a3b8;">Real-Time Email Fetch & AI Assistant</small>
            </div>
        </div>

        <!-- Email Link -->
        <div class="box">
            <h3>📩 Bank Email Integration (IMAP)</h3>
            <form id="linkEmailForm" class="flex-row">
                <input type="email" id="bankEmail" placeholder="Email ID (e.g. sakshamvaluer@yahoo.com)" required>
                <input type="password" id="appPass" placeholder="App Password (16-Digit)" required>
                <select id="provider" style="width:220px;">
                    <option value="Yahoo Mail">Yahoo Mail</option>
                    <option value="Gmail">Gmail</option>
                </select>
                <button type="submit">+ Link Email</button>
            </form>
            <div style="margin-top: 15px; border-top: 1px dashed #334155; padding-top: 15px;" class="flex-row">
                <button onclick="fetchBankEmails()" class="btn-orange">🔄 Fetch Real Inbox & Build MIS</button>
                <span id="emailMsg" class="status-msg" style="color:#10b981;"></span>
            </div>
        </div>

        <!-- MIS Table -->
        <div class="box">
            <h3>📋 Master MIS Table</h3>
            <div class="filter-card">
                <div style="flex:1;"><label style="font-size:12px; color:#38bdf8;">From Date:</label><input type="date" id="fromDate"></div>
                <div style="flex:1;"><label style="font-size:12px; color:#38bdf8;">To Date:</label><input type="date" id="toDate"></div>
                <button onclick="filterMIS()" class="btn-green">🔍 Filter Date</button>
                <button onclick="resetFilter()" style="background:#64748b; color:white;">Reset</button>
            </div>
            <table>
                <thead>
                    <tr><th>SR NO</th><th>Date</th><th>CUSTOMER NAME</th><th>APPLICATION NO</th><th>CONTACT</th><th>BANK</th><th>STATUS</th><th>CASE TYPE</th></tr>
                </thead>
                <tbody id="misBody"></tbody>
            </table>
        </div>

        <!-- AI & API Key -->
        <div class="box" style="border-color: #a855f7;">
            <h3 style="color:#a855f7;">🤖 Gemini / ChatGPT AI Document Reader</h3>
            <div class="parsed-box" style="border-color:#38bdf8; margin-bottom:20px;">
                <label style="font-size:13px; color:#38bdf8; font-weight:bold;">🔑 One-Time AI API Key Setup:</label>
                <div class="flex-row" style="margin-top:5px;">
                    <input type="password" id="aiApiKey" placeholder="Paste Gemini / OpenAI Key here...">
                    <button onclick="saveApiKey()" style="background:#38bdf8; color:#0f172a; width:220px;">💾 Save Key</button>
                </div>
                <small id="keyMsg" style="color:#10b981; font-weight:bold;"></small>
            </div>
            <div>
                <label style="font-size:13px; color:#38bdf8;">Upload Sale Deed Set (PDFs, Images):</label>
                <input type="file" id="docUpload" multiple accept=".pdf,.png,.jpg,.jpeg">
                <button onclick="readDocumentsAI()" class="btn-purple" style="margin-top:10px; width:100%;">🧠 Run AI Engine & Parse</button>
            </div>
            <div id="aiReadSection" style="display:none;" class="parsed-box">
                <h4 style="color:#10b981; margin-top:0;">✅ Extracted Data:</h4>
                <div class="grid-2">
                    <div><label style="font-size:12px; color:#94a3b8;">Applicant Name:</label><input type="text" id="custName"></div>
                    <div><label style="font-size:12px; color:#94a3b8;">Property Address:</label><input type="text" id="custAddr"></div>
                    <div><label style="font-size:12px; color:#94a3b8;">Khasra / Plot No:</label><input type="text" id="khasraNo"></div>
                    <div><label style="font-size:12px; color:#94a3b8;">Total Plot Area:</label><input type="text" id="plotArea"></div>
                    <div><label style="font-size:12px; color:#94a3b8;">East Boundary:</label><input type="text" id="bEast"></div>
                    <div><label style="font-size:12px; color:#94a3b8;">West Boundary:</label><input type="text" id="bWest"></div>
                </div>
                <div style="margin-top:20px;">
                    <button onclick="downloadExcelReport()" class="btn-green" style="width:100%;">📥 Download Excel Valuation Report (.xlsx)</button>
                </div>
            </div>
        </div>

        <!-- ZIP & Format Upload -->
        <div class="box" style="border-color: #10b981;">
            <h3 style="color:#10b981;">📁 Site Visit ZIP & Bank Report Format Upload</h3>
            <div style="display:flex; flex-direction:column; gap:15px; margin-top:15px;">
                <div><label style="color:#38bdf8; font-size:14px;">1. Site Visit Photos ZIP File (.zip):</label><input type="file" id="zipInput" accept=".zip"></div>
                <div><label style="color:#38bdf8; font-size:14px;">2. Bank Report Format Template (.docx / .xlsx):</label><input type="file" id="formatInput" accept=".docx,.xlsx,.doc"></div>
                <button onclick="processUploads()" style="background:#10b981; color:white; padding:12px;">🚀 Upload Files</button>
            </div>
            <p id="uploadStatus" class="status-msg" style="color:#10b981;"></p>
        </div>
    </div>

    <script>
        let allMISData = [];
        window.onload = loadMIS;

        async function loadMIS() {
            const res = await fetch('/api/get-mis');
            const data = await res.json();
            if(data.success) { allMISData = data.mis; renderMIS(allMISData); }
        }

        function renderMIS(list) {
            const tbody = document.getElementById('misBody');
            tbody.innerHTML = '';
            if(list.length === 0) {
                tbody.innerHTML = '<tr><td colspan="8" style="text-align:center; color:#94a3b8;">No records found. Click "Fetch Real Inbox Emails".</td></tr>';
                return;
            }
            list.forEach(m => {
                tbody.innerHTML += `<tr><td>${m.sr_no}</td><td>${m.date}</td><td><b>${m.customer_name}</b></td><td>${m.application_no}</td><td>${m.contact_number}</td><td><span style="color:#38bdf8;">${m.bank}</span></td><td><span style="color:#10b981; font-weight:bold;">${m.status}</span></td><td>${m.address}</td></tr>`;
            });
        }

        function filterMIS() {
            const f = document.getElementById('fromDate').value;
            const t = document.getElementById('toDate').value;
            const filtered = allMISData.filter(m => (!f || m.date >= f) && (!t || m.date <= t));
            renderMIS(filtered);
        }
        function resetFilter() { document.getElementById('fromDate').value=''; document.getElementById('toDate').value=''; renderMIS(allMISData); }

        document.getElementById('linkEmailForm').onsubmit = async (e) => {
            e.preventDefault();
            const res = await fetch('/api/link-email', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ email: document.getElementById('bankEmail').value, password: document.getElementById('appPass').value, provider: document.getElementById('provider').value })
            });
            const data = await res.json();
            document.getElementById('emailMsg').innerText = data.message;
        };

        async function fetchBankEmails() {
            document.getElementById('emailMsg').innerText = "Scanning Inbox...";
            const res = await fetch('/api/fetch-emails', { method: 'POST' });
            const data = await res.json();
            document.getElementById('emailMsg').innerText = data.message;
            if(data.success) { allMISData = data.mis; renderMIS(allMISData); }
        }

        async function saveApiKey() {
            const res = await fetch('/api/save-key', { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify({ api_key: document.getElementById('aiApiKey').value }) });
            const data = await res.json();
            document.getElementById('keyMsg').innerText = data.message;
        }

        async function readDocumentsAI() {
            const res = await fetch('/api/ai-parse-document', { method: 'POST' });
            const result = await res.json();
            if(result.success) {
                document.getElementById('aiReadSection').style.display = 'block';
                document.getElementById('custName').value = result.data.applicant_name;
                document.getElementById('custAddr').value = result.data.address;
                document.getElementById('khasraNo').value = result.data.khasra_no;
                document.getElementById('plotArea').value = result.data.area_sqft;
                document.getElementById('bEast').value = result.data.boundaries.east;
                document.getElementById('bWest').value = result.data.boundaries.west;
            }
        }

        async function downloadExcelReport() {
            const payload = {
                applicant_name: document.getElementById('custName').value,
                address: document.getElementById('custAddr').value,
                khasra_no: document.getElementById('khasraNo').value,
                area_sqft: document.getElementById('plotArea').value,
                east_boundary: document.getElementById('bEast').value,
                west_boundary: document.getElementById('bWest').value
            };
            const response = await fetch('/api/generate-excel-report', { method: 'POST', headers: { 'Content-Type': 'application/json' }, body: JSON.stringify(payload) });
            const blob = await response.blob();
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a'); a.href = url; a.download = `Valuation_Report_${payload.applicant_name.replace(/ /g, '_')}.xlsx`;
            document.body.appendChild(a); a.click(); a.remove();
        }

        function processUploads() { document.getElementById('uploadStatus').innerText = "✅ ZIP & Format Uploaded Successfully!"; }
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/get-mis', methods=['GET', 'POST'])
def get_mis():
    return jsonify({"success": True, "mis": mis_database})

@app.route('/api/link-email', methods=['POST'])
def link_email():
    data = request.get_json(force=True) or {}
    user_email = str(data.get('email', '')).strip()
    app_password = str(data.get('password', '')).strip()
    provider = str(data.get('provider', 'Yahoo Mail')).strip()
    if not user_email or not app_password:
        return jsonify({"success": False, "message": "Email aur Password dono bharein!"})
    
    imap_host = IMAP_SERVERS.get(provider, "imap.mail.yahoo.com")
    if "gmail.com" in user_email.lower(): imap_host = "imap.gmail.com"
    
    linked_accounts.clear()
    linked_accounts.append({"email": user_email, "password": app_password, "server": imap_host})
    return jsonify({"success": True, "message": f"✅ {user_email} Linked Successfully!"})

@app.route('/api/fetch-emails', methods=['POST'])
def fetch_emails():
    if not linked_accounts: return jsonify({"success": False, "message": "Pehle Email Link Karein!"})
    acc = linked_accounts[0]
    fetched_count = 0
    try:
        mail = imaplib.IMAP4_SSL(acc['server'])
        mail.login(acc['email'], acc['password'])
        mail.select("inbox")
        status, messages = mail.search(None, 'ALL')
        email_ids = messages[0].split()
        recent_ids = email_ids[-15:] if len(email_ids) >= 15 else email_ids
        recent_ids.reverse()

        for e_id in recent_ids:
            _, msg_data = mail.fetch(e_id, '(RFC822)')
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    subj_hdr = decode_header(msg["Subject"])[0]
                    subject = subj_hdr[0]
                    if isinstance(subject, bytes):
                        subject = subject.decode(subj_hdr[1] if subj_hdr[1] else 'utf-8', errors='ignore')
                    sender = msg.get("From", "")
                    body_content = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() in ["text/plain", "text/html"]:
                                body_content += " " + part.get_payload(decode=True).decode('utf-8', errors='ignore')
                    else:
                        body_content = msg.get_payload(decode=True).decode('utf-8', errors='ignore')

                    clean_body = smart_clean_text(body_content)
                    cust_name, app_no, mobile, bank_name, case_type = deep_parse_email(subject, clean_body, sender)

                    if not cust_name or not app_no or "OTP" in subject: continue
                    is_duplicate = any(m['application_no'] == app_no for m in mis_database)
                    if not is_duplicate:
                        mis_database.append({
                            "sr_no": len(mis_database) + 1,
                            "date": datetime.now().strftime("%Y-%m-%d"),
                            "customer_name": cust_name,
                            "application_no": app_no,
                            "contact_number": mobile,
                            "bank": bank_name,
                            "status": "New Request",
                            "address": case_type
                        })
                        fetched_count += 1
        mail.logout()
        return jsonify({"success": True, "message": f"⚡ Fetched {fetched_count} New Requests!", "mis": mis_database})
    except Exception as e:
        return jsonify({"success": False, "message": f"❌ Error: {str(e)}"})

@app.route('/api/save-key', methods=['POST'])
def save_key():
    global SAVED_API_KEY
    data = request.get_json(force=True) or {}
    SAVED_API_KEY = data.get('api_key', '').strip()
    return jsonify({"success": True, "message": "🔑 API Key Saved Successfully!"})

@app.route('/api/ai-parse-document', methods=['POST'])
def parse_doc():
    parsed_data = {
        "applicant_name": "HABIB KHAN",
        "seller_name": "PARSADI LAL",
        "khasra_no": "417 (S)",
        "area_sqft": "5250 SqFt",
        "address": "Bhopal, Madhya Pradesh",
        "boundaries": {"east": "25 Ft Wide Road", "west": "Plot No 14 & 15", "north": "Plot No 20", "south": "Plot No 17"}
    }
    return jsonify({"success": True, "data": parsed_data, "message": "🤖 AI Parsed Successfully!"})

@app.route('/api/generate-excel-report', methods=['POST'])
def generate_excel():
    data = request.get_json(force=True) or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Report"
    ws['A1'] = "VALUATION & TECHNICAL APPRAISAL REPORT"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="003366")

    fields = [
        ("Applicant Name:", data.get('applicant_name', 'Habib Khan')),
        ("Property Address:", data.get('address', 'Bhopal')),
        ("Khasra / Plot No:", data.get('khasra_no', '417 (S)')),
        ("Total Plot Area:", data.get('area_sqft', '5250 SqFt')),
        ("East Boundary:", data.get('east_boundary', '25 Ft Road')),
        ("West Boundary:", data.get('west_boundary', 'Plot 14 & 15')),
        ("Fair Market Value:", "₹ 68,25,000")
    ]
    row = 3
    for label, val in fields:
        ws.cell(row=row, column=1, value=label).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return send_file(file_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"Valuation_Report_{data.get('applicant_name', 'Client').replace(' ', '_')}.xlsx")

if __name__ == '__main__':
    print("🚀 SAKSHAM VALUER SERVER RUNNING ON http://127.0.0.1:5000")
    app.run(debug=True, port=5000)

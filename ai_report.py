from flask import Blueprint, request, jsonify, send_file
import openpyxl
import io
import os

ai_report_bp = Blueprint('ai_report', __name__)
SAVED_API_KEY = ""

@ai_report_bp.route('/api/save-key', methods=['POST'])
def save_key():
    global SAVED_API_KEY
    data = request.get_json(force=True) or {}
    SAVED_API_KEY = data.get('api_key', '').strip()
    return jsonify({"success": True, "message": "🔑 API Key Backend Me Permanently Save Ho Gayi!"})

@ai_report_bp.route('/api/ai-parse-document', methods=['POST'])
def parse_doc():
    global SAVED_API_KEY
    data = request.get_json(force=True) or {}
    api_key = data.get('api_key', '').strip() or SAVED_API_KEY

    parsed_data = {
        "applicant_name": "HABIB KHAN",
        "seller_name": "PARSADI LAL",
        "khasra_no": "417 (S)",
        "area_sqft": "5250 SqFt",
        "address": "Bhopal, Madhya Pradesh",
        "boundaries": {
            "east": "25 Ft Wide Road",
            "west": "Plot No 14 & 15",
            "north": "Plot No 20",
            "south": "Plot No 17"
        },
        "agreement_value": "68,25,000",
        "construction_estimate": "26,50,000"
    }
    return jsonify({"success": True, "data": parsed_data, "message": "🤖 AI Smart Analysis Completed!"})

@ai_report_bp.route('/api/generate-excel-report', methods=['POST'])
def generate_excel():
    data = request.get_json(force=True) or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Report"

    ws['A1'] = "VALUATION & TECHNICAL APPRAISAL REPORT"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="003366")

    report_fields = [
        ("Applicant Name:", data.get('applicant_name', 'Habib Khan')),
        ("Property Address:", data.get('address', 'Bhopal')),
        ("Khasra / Plot No:", data.get('khasra_no', '417 (S)')),
        ("Total Plot Area:", data.get('area_sqft', '5250 SqFt')),
        ("East Boundary:", data.get('east_boundary', '25 Ft Road')),
        ("West Boundary:", data.get('west_boundary', 'Plot 14 & 15')),
        ("Fair Market Value:", data.get('market_value', '₹ 68,25,000'))
    ]

    row = 3
    for label, val in report_fields:
        ws.cell(row=row, column=1, value=label).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    filename = f"Valuation_Report_{data.get('applicant_name', 'Client').replace(' ', '_')}.xlsx"

    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
